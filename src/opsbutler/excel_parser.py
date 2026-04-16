from openpyxl import load_workbook
from datetime import datetime
from opsbutler.models import SheetData, ExcelPayload, ExcelSummary, ScheduleTable
from typing import Any


def load_excel(file_path: str, config=None) -> ExcelPayload:
    """
    Main entry: reads an Excel file, parses all sheets dynamically,
    detects action columns, and returns a complete ExcelPayload.

    Args:
        file_path: Path to the Excel file
        config: Config object with excel.action_column_candidates etc.
                If None, uses defaults.
    """
    # Get candidates from config or use defaults
    action_candidates = ["操作类型", "操作", "变更事项", "变更类型", "任务类型"]
    app_candidates = ["APPID", "应用", "应用名称", "应用名"]
    skip_sheets = ["变更安排"]

    if config:
        action_candidates = config.excel.action_column_candidates
        app_candidates = config.excel.app_column_candidates
        skip_sheets = config.excel.skip_sheets

    wb = load_workbook(file_path, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        if ws.max_row <= 1:  # skip empty sheets (header only)
            continue
        sheet_data = _parse_sheet(ws, action_candidates, app_candidates)
        sheets.append(sheet_data)

    summary = _build_summary(sheets)
    return ExcelPayload(
        source_file=file_path,
        sheets=sheets,
        summary=summary,
    )


def load_schedule_sheet(file_path: str) -> ScheduleTable | None:
    """Parse the '变更安排' schedule sheet separately.
    Returns None if the sheet does not exist or is empty.
    """
    wb = load_workbook(file_path, data_only=True)
    sheet_name = "变更安排"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    if ws.max_row <= 1:
        return None

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[header] = _serialize_value(cell.value)
        rows.append(row_data)

    return ScheduleTable(headers=headers, rows=rows)


def _parse_sheet(ws, action_candidates: list[str], app_candidates: list[str]) -> SheetData:
    """Parse a single worksheet into SheetData."""
    headers = [cell.value for cell in ws[1]]
    # Clean headers: strip whitespace, convert None to empty string
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[header] = _serialize_value(cell.value)
        rows.append(row_data)

    detected_action = _detect_column(headers, action_candidates)
    detected_app = _detect_column(headers, app_candidates)

    return SheetData(
        sheet_name=ws.title,
        headers=headers,
        rows=rows,
        detected_action_column=detected_action,
        detected_app_column=detected_app,
    )


def _detect_column(headers: list[str], candidates: list[str]) -> str | None:
    """
    Find a matching column from headers using candidate names.
    Strategy: exact match first, then contains match.
    """
    # Exact match
    for candidate in candidates:
        if candidate in headers:
            return candidate
    # Contains match (candidate is substring of a header)
    for candidate in candidates:
        for header in headers:
            if candidate in header:
                return header
    return None


def _serialize_value(val) -> Any:
    """Convert cell value to JSON-safe format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)


def _build_summary(sheets: list[SheetData]) -> ExcelSummary:
    """Aggregate metadata across all sheets."""
    all_apps = set()
    all_op_types = set()
    total_rows = 0

    for sheet in sheets:
        total_rows += len(sheet.rows)
        # Collect unique apps
        if sheet.detected_app_column:
            for row in sheet.rows:
                app = row.get(sheet.detected_app_column)
                if app:
                    all_apps.add(str(app))
        # Collect unique operation types
        if sheet.detected_action_column:
            for row in sheet.rows:
                op = row.get(sheet.detected_action_column)
                if op:
                    all_op_types.add(str(op))

    return ExcelSummary(
        total_sheets=len(sheets),
        total_rows=total_rows,
        unique_apps=sorted(all_apps),
        unique_operation_types=sorted(all_op_types),
        sheet_names=[s.sheet_name for s in sheets],
    )
